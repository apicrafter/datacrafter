"""Tests for source classes"""
import json
import os
import zipfile

import pytest

from datacrafter.sources.bsonf import BSONSource
from datacrafter.sources.csv import CSVSource
from datacrafter.sources.json import JSONSource
from datacrafter.sources.jsonl import JSONLinesSource
from datacrafter.sources.xml import XMLSource
from datacrafter.sources.xlsx import XLSXSource
from datacrafter.sources.xls import XLSSource
from datacrafter.sources.zipped import ZIPSourceWrapper
from datacrafter.sources.zipxml import ZIPXMLSource


class TestJSONLinesSource:
    """Tests for JSONLinesSource"""
    
    def test_read_single_record(self, jsonl_file):
        """Test reading a single record"""
        source = JSONLinesSource(filename=jsonl_file)
        record = source.read()
        
        assert record['id'] == 1
        assert record['name'] == 'Alice'
        assert record['age'] == 30
    
    def test_read_multiple_records(self, jsonl_file):
        """Test reading multiple records"""
        source = JSONLinesSource(filename=jsonl_file)
        records = []
        try:
            for _ in range(3):
                records.append(source.read())
        except (StopIteration, ValueError):
            pass  # End of file or parsing error
        
        assert len(records) == 3
        assert records[0]['name'] == 'Alice'
        assert records[1]['name'] == 'Bob'
        assert records[2]['name'] == 'Charlie'
    
    def test_read_bulk(self, jsonl_file):
        """Test reading bulk records"""
        source = JSONLinesSource(filename=jsonl_file)
        records = source.read_bulk(2)
        
        assert len(records) == 2
        assert records[0]['id'] == 1
        assert records[1]['id'] == 2
    
    def test_reset(self, jsonl_file):
        """Test resetting the source"""
        source = JSONLinesSource(filename=jsonl_file)
        record1 = source.read()
        source.reset()
        record2 = source.read()
        
        assert record1 == record2
    
    def test_context_manager(self, jsonl_file):
        """Test using source as context manager"""
        with JSONLinesSource(filename=jsonl_file) as source:
            record = source.read()
            assert record is not None
        # File should be closed after context


class TestCSVSource:
    """Tests for CSVSource"""
    
    def test_read_single_record(self, csv_file):
        """Test reading a single CSV record"""
        source = CSVSource(filename=csv_file, keys=None)
        record = source.read()
        
        assert record['id'] == '1'
        assert record['name'] == 'Alice'
        assert record['age'] == '30'
    
    def test_read_with_custom_keys(self, csv_file):
        """Test reading CSV with custom keys"""
        source = CSVSource(filename=csv_file, keys=['id', 'name', 'age'])
        record = source.read()
        
        assert 'id' in record
        assert 'name' in record
        assert 'age' in record
    
    def test_is_flat(self, csv_file):
        """Test that CSV source is flat"""
        source = CSVSource(filename=csv_file)
        assert source.is_flat() is True


class TestJSONSource:
    def test_read_array_and_bulk(self, temp_dir):
        path = os.path.join(temp_dir, 'data.json')
        with open(path, 'w', encoding='utf8') as file_obj:
            json.dump([{'id': 1}, {'id': 2}, {'id': 3}], file_obj)
        source = JSONSource(filename=path)
        assert source.id() == 'json'
        assert source.read() == {'id': 1}
        assert source.read_bulk(2) == [{'id': 2}, {'id': 3}]
        with pytest.raises(StopIteration):
            source.read()

    def test_tagname_selects_nested_list(self, temp_dir):
        path = os.path.join(temp_dir, 'nested.json')
        with open(path, 'w', encoding='utf8') as file_obj:
            json.dump({'records': [{'name': 'Ada'}]}, file_obj)
        source = JSONSource(filename=path, tagname='records')
        assert source.read() == {'name': 'Ada'}


class TestBSONSource:
    def test_read_and_bulk(self, temp_dir):
        bson = pytest.importorskip('bson')
        path = os.path.join(temp_dir, 'data.bson')
        with open(path, 'wb') as file_obj:
            file_obj.write(bson.encode({'id': 1}))
            file_obj.write(bson.encode({'id': 2}))
            file_obj.write(bson.encode({'id': 3}))
        source = BSONSource(filename=path)
        assert source.id() == 'bson'
        assert source.read()['id'] == 1
        assert [doc['id'] for doc in source.read_bulk(2)] == [2, 3]


class TestXMLSource:
    def test_read_tagged_elements(self, temp_dir):
        pytest.importorskip('lxml.etree')
        path = os.path.join(temp_dir, 'data.xml')
        with open(path, 'wb') as file_obj:
            file_obj.write(
                b'<root><item><name>Ada</name></item>'
                b'<item><name>Bob</name></item></root>')
        source = XMLSource(filename=path, tagname='item')
        assert source.id() == 'xml'
        assert source.is_flat() is False
        assert source.read() == {'name': 'Ada'}
        assert source.read_bulk(1) == [{'name': 'Bob'}]


class TestXLSXSource:
    def test_read_rows_skipping_header(self, temp_dir):
        pytest.importorskip('openpyxl')
        from openpyxl import Workbook

        path = os.path.join(temp_dir, 'data.xlsx')
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['id', 'name'])
        sheet.append([1, 'Ada'])
        sheet.append([2, 'Bob'])
        workbook.save(path)
        source = XLSXSource(
            filename=path, keys=['id', 'name'], start_line=2)
        assert source.id() == 'xlsx'
        assert source.is_flat() is True
        assert source.read() == {'id': '1', 'name': 'Ada'}
        assert source.read_bulk(1) == [{'id': '2', 'name': 'Bob'}]

    def test_reads_selected_sheet(self, temp_dir):
        pytest.importorskip('openpyxl')
        from openpyxl import Workbook

        path = os.path.join(temp_dir, 'paged.xlsx')
        workbook = Workbook()
        first = workbook.active
        first.append(['id', 'name'])
        first.append([1, 'Ada'])
        second = workbook.create_sheet('other')
        second.append(['id', 'name'])
        second.append([9, 'Zoe'])
        workbook.save(path)
        source = XLSXSource(
            filename=path, keys=['id', 'name'], page=1, start_line=2)
        assert source.read() == {'id': '9', 'name': 'Zoe'}


class TestXLSSource:
    def test_read_rows_skipping_header(self, temp_dir):
        pytest.importorskip('xlrd')
        pandas = pytest.importorskip('pandas')
        path = os.path.join(temp_dir, 'data.xls')
        try:
            pandas.DataFrame(
                {'id': [1, 2], 'name': ['Ada', 'Bob']}
            ).to_excel(path, index=False, engine='xlwt')
        except (ImportError, ValueError, ModuleNotFoundError):
            pytest.skip('xlwt is not available to write .xls fixtures')
        source = XLSSource(
            filename=path, keys=['id', 'name'], start_line=1)
        assert source.id() == 'xls'
        assert source.is_flat() is True
        row = source.read()
        assert row['name'] == 'Ada'
        bulk = source.read_bulk(1)
        assert bulk[0]['name'] == 'Bob'


class TestZIPSourceWrapper:
    def test_iterates_files_in_archive(self, temp_dir):
        path = os.path.join(temp_dir, 'bundle.zip')
        with zipfile.ZipFile(path, 'w') as archive:
            archive.writestr('a.txt', 'one\n')
            archive.writestr('b.txt', 'two\n')

        class LineZip(ZIPSourceWrapper):
            def read_single(self):
                line = self.current_file.readline()
                if not line:
                    raise StopIteration
                if isinstance(line, bytes):
                    line = line.decode('utf8')
                return line.strip()

        source = LineZip(path)
        assert source.read() == 'one'
        assert source.read() == 'two'
        with pytest.raises(StopIteration):
            source.read()
        source.close()


class TestZIPXMLSource:
    def test_reads_tagged_elements_across_files(self, temp_dir):
        pytest.importorskip('lxml')
        path = os.path.join(temp_dir, 'bundle.zip')
        xml_a = (
            '<?xml version="1.0"?><root>'
            '<item><name>Ada</name></item>'
            '</root>'
        )
        xml_b = (
            '<?xml version="1.0"?><root>'
            '<item><name>Bob</name></item>'
            '</root>'
        )
        with zipfile.ZipFile(path, 'w') as archive:
            archive.writestr('a.xml', xml_a)
            archive.writestr('b.xml', xml_b)
        source = ZIPXMLSource(filename=path, tagname='item')
        assert source.id() == 'zip-xml'
        assert source.is_flat() is False
        first = source.read()
        assert first['name'] == 'Ada'
        rest = source.read_bulk(5)
        assert rest[0]['name'] == 'Bob'
        source.reset()
        assert source.read()['name'] == 'Ada'
        source.close()

    def test_factory_builds_zipxml(self, temp_dir):
        pytest.importorskip('lxml')
        from datacrafter.sources import get_source_from_file
        path = os.path.join(temp_dir, 'one.zip')
        with zipfile.ZipFile(path, 'w') as archive:
            archive.writestr(
                'a.xml',
                '<?xml version="1.0"?><root><row><id>1</id></row></root>')
        source = get_source_from_file(
            path, stype='zipxml', options={'tagname': 'row'})
        assert isinstance(source, ZIPXMLSource)
        assert source.read()['id'] == '1'
        source.close()


class TestXLSSourceMocked:
    def test_read_single_row_types(self, monkeypatch):
        from datacrafter.sources import xls as xls_mod
        if not xls_mod.HAS_XLRD:
            pytest.skip('xlrd is not installed')

        class FakeSheet:
            ncols = 3

            def cell_type(self, _row, col):
                return [
                    xls_mod.xlrd.XL_CELL_NUMBER,
                    xls_mod.xlrd.XL_CELL_DATE,
                    xls_mod.xlrd.XL_CELL_TEXT,
                ][col]

            def cell_value(self, _row, col):
                return [7.0, 44927.0, 'Ada'][col]

        monkeypatch.setattr(
            xls_mod.xlrd, 'xldate_as_tuple',
            lambda _value, _datemode: (2023, 1, 2, 0, 0, 0))
        row = xls_mod.read_single_row(
            0, 3, 0, ['n', 'd', 's'], FakeSheet())
        assert row['n'] == 7
        assert row['s'] == 'Ada'
        assert row['d'].startswith('2023-01-02')

    def test_xls_source_reads_mocked_workbook(self, temp_dir, monkeypatch):
        from datacrafter.sources import xls as xls_mod
        if not xls_mod.HAS_XLRD:
            pytest.skip('xlrd is not installed')

        class FakeSheet:
            ncols = 2
            nrows = 2

            def cell_type(self, _row, _col):
                return xls_mod.xlrd.XL_CELL_TEXT

            def cell_value(self, row, col):
                return [['1', 'Ada'], ['2', 'Bob']][row][col]

        class FakeBook:
            datemode = 0

            def sheet_by_index(self, _page):
                return FakeSheet()

        monkeypatch.setattr(xls_mod, 'open_workbook', lambda _name: FakeBook())
        path = os.path.join(temp_dir, 'data.xls')
        with open(path, 'wb') as file_obj:
            file_obj.write(b'placeholder')
        source = XLSSource(
            filename=path, keys=['id', 'name'], start_line=0)
        assert source.id() == 'xls'
        assert source.is_flat() is True
        assert source.read() == {'id': '1', 'name': 'Ada'}
        assert source.read_bulk(1) == [{'id': '2', 'name': 'Bob'}]
        with pytest.raises(StopIteration):
            source.read()
        source.close()

