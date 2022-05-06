from openpyxl import load_workbook
from .base import BaseFileSource

class XLSXSource(BaseFileSource):
    def __init__(self, filename=None, stream=None, keys=None, page=0, start_line=1):
        super(XLSXSource, self).__init__(filename, stream, binary=False, noopen=True)
        self.keys = keys
        self.start_line = start_line
        self.page = page
        self.pos = self.start_line
        self.reset()
        pass

    def reset(self):
        super(XLSXSource, self).reset()
        self.workbook = load_workbook(self.filename)
        self.sheet = self.workbook.active
        self.pos = self.start_line
        self.iter = self.sheet.iter_rows()

    def id(self):
        return 'xlsx'

    def is_flat(self):
        return True


    def read(self):
        """Read single XLSX record"""
        row = next(self.iter)
        tmp = list()
        for cell in row:
            tmp.append(cell.value)
        result = dict(zip(self.keys, tmp))
        self.pos += 1
        return result

    def read_bulk(self, num):
        """Read bulk XLSX records"""
        chunk = []
        for n in range(0, num):
            row = next(self.iter)
            tmp = list()
            for cell in row:
                tmp.append(cell.value)
            result = dict(zip(self.keys, tmp))
            chunk.append(result)
            self.pos += 1
        return chunk
