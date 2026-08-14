"""XLSX source module."""
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    load_workbook = None

from .._registry import register_source
from .base import BaseFileSource


@register_source("xlsx")
class XLSXSource(BaseFileSource):
    """XLSX source implementation."""
    def __init__(self, filename=None, stream=None, keys=None, page=0, start_line=1):
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for XLSXSource. "
                "Install it with: pip install openpyxl"
            )
        super().__init__(filename, stream, binary=False, noopen=True)
        self.keys = keys
        self.start_line = start_line
        self.page = page
        self.pos = self.start_line
        self.reset()

    def reset(self):
        super().reset()
        self.workbook = load_workbook(self.filename)
        if isinstance(self.page, str):
            self.sheet = self.workbook[self.page]
        else:
            self.sheet = self.workbook.worksheets[self.page]
        self.pos = self.start_line
        self.iter = self.sheet.iter_rows()
        if self.pos > 1:
            self.skip(self.pos - 1)

    def skip(self, num):
        """Skip num rows."""
        while num > 0:
            num -= 1
            next(self.iter)

    def id(self):
        return 'xlsx'

    def is_flat(self):
        return True

    def read(self):
        """Read single XLSX record"""
        row = next(self.iter)
        tmp = []
        for cell in row:
            tmp.append(str(cell.value))
        result = dict(zip(self.keys, tmp))
        self.pos += 1
        return result

    def read_bulk(self, num):
        """Read bulk XLSX records"""
        chunk = []
        for _ in range(0, num):
            row = next(self.iter)
            tmp = []
            for cell in row:
                tmp.append(str(cell.value))
            result = dict(zip(self.keys, tmp))
            chunk.append(result)
            self.pos += 1
        return chunk
